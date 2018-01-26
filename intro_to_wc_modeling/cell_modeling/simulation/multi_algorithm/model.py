'''
Reads models specified in Excel into a Python object

@author Jonathan Karr, karr@mssm.edu
@date 3/22/2016
'''

# required libraries
from cobra import Metabolite as CobraMetabolite
from cobra import Model as CobraModel
from cobra import Reaction as CobraReaction
from intro_to_wc_modeling.cell_modeling.simulation.multi_algorithm import util
from intro_to_wc_modeling.cell_modeling.simulation.multi_algorithm.util import N_AVOGADRO
from itertools import chain
from numpy import random
from openpyxl import load_workbook
import math
import numpy as np
import re
import warnings


class Model(object):
    # Represents a model (submodels, compartments, species, reactions, parameters, references)

    submodels = []
    compartments = []
    species = []
    reactions = []
    parameters = []
    references = []

    density = None
    fractionDryWeight = None

    speciesCounts = np.zeros(0)  # rows: species, columns: compartments
    mass = None  # cell mass
    dryWeight = None  # cell dry weight
    volume = None  # cell volume
    extracellularVolume = None  # media volume
    growth = None

    def __init__(self, submodels=None, compartments=None, species=None, reactions=None, parameters=None, references=None):
        if submodels is None:
            submodels = []
        if compartments is None:
            compartments = []
        if species is None:
            species = []
        if reactions is None:
            reactions = []
        if parameters is None:
            parameters = []
        if references is None:
            references = []
        self.submodels = submodels
        self.compartments = compartments
        self.species = species
        self.reactions = reactions
        self.parameters = parameters
        self.references = references

    '''
    def __init__(self):
        self.submodels = []
        self.compartments = []
        self.species = []
        self.reactions = []
        self.parameters = []
        self.references = []
    '''

    def setupSimulation(self):
        self.fractionDryWeight = self.getComponentById('fractionDryWeight', self.parameters).value

        for subModel in self.submodels:
            subModel.setupSimulation()

        self.calcInitialConditions()

    def calcInitialConditions(self):
        cellComp = self.getComponentById('c', self.compartments)
        extrComp = self.getComponentById('e', self.compartments)

        # volume
        self.volume = cellComp.initialVolume
        self.extracellularVolume = extrComp.initialVolume

        # species counts
        self.speciesCounts = np.zeros((len(self.species), len(self.compartments)))
        for species in self.species:
            for conc in species.concentrations:
                self.speciesCounts[species.index, conc.compartment.index] = conc.value * conc.compartment.initialVolume * N_AVOGADRO

        # cell mass
        self.calcMass()

        # density
        self.density = self.mass / self.volume

        # growth
        self.growth = np.nan

        # sync submodels
        for subModel in self.submodels:
            subModel.updateLocalCellState(self)

    def calcMass(self):
        for comp in self.compartments:
            if comp.id == 'c':
                iCellComp = comp.index

        mass = 0.
        for species in self.species:
            if species.molecularWeight is not None:
                mass += self.speciesCounts[species.index, iCellComp] * species.molecularWeight
        mass /= N_AVOGADRO

        self.mass = mass
        self.dryWeight = self.fractionDryWeight * mass

    def calcVolume(self):
        self.volume = self.mass / self.density

    def setComponentIndices(self):
        for index, obj in enumerate(self.submodels):
            obj.index = index
        for index, obj in enumerate(self.compartments):
            obj.index = index
        for index, obj in enumerate(self.species):
            obj.index = index
        for index, obj in enumerate(self.reactions):
            obj.index = index
        for index, obj in enumerate(self.parameters):
            obj.index = index
        for index, obj in enumerate(self.references):
            obj.index = index

    def getSpeciesCountsDict(self):
        # get species counts as dictionary

        speciesCountsDict = {}
        for species in self.species:
            for compartment in self.compartments:
                speciesCountsDict['%s[%s]' % (species.id, compartment.id)] = self.speciesCounts[species.index, compartment.index]
        return speciesCountsDict

    def setSpeciesCountsDict(self, speciesCountsDict):
        # set species counts for dictionary

        for species in self.species:
            for compartment in self.compartments:
                self.speciesCounts[species.index, compartment.index] = speciesCountsDict['%s[%s]' % (species.id, compartment.id)]

    def getComponentById(self, id, components=None):
        if not components:
            components = chain(self.submodels, self.compartments, self.species, self.reactions, self.parameters, self.references)

        for component in components:
            if component.id == id:
                return component


class Submodel(object):
    # Represents a model (submodels, compartments, species, reactions, parameters, references)

    index = None
    id = ''
    name = ''
    algorithm = ''

    reactions = []
    species = []
    parameters = []

    speciesCounts = np.zeros(0)
    volume = np.zeros(0)
    extracellularVolume = np.zeros(0)

    # fix
    def __init__(self, id='', name='', reactions=[], species=[]):
        self.id = id
        self.name = name
        self.reactions = reactions
        self.species = species

    def setupSimulation(self):
        # initialize species counts dictionary
        self.speciesCounts = {}
        for species in self.species:
            self.speciesCounts[species.id] = 0

    def updateLocalCellState(self, model):
        # sets local species counts from global species counts
        for species in self.species:
            self.speciesCounts[species.id] = model.speciesCounts[species.species.index, species.compartment.index]
        self.volume = model.volume
        self.extracellularVolume = model.extracellularVolume

    def updateGlobalCellState(self, model):
        # sets global species counts from local species counts
        for species in self.species:
            model.speciesCounts[species.species.index, species.compartment.index] = self.speciesCounts[species.id]

    def getSpeciesConcentrations(self):
        # get species concentrations
        volumes = self.getSpeciesVolumes()
        concs = {}
        for species in self.species:
            concs[species.id] = self.speciesCounts[species.id] / volumes[species.id] / N_AVOGADRO
        return concs

    def getSpeciesVolumes(self):
        # get container volumes for each species
        volumes = {}
        for species in self.species:
            if species.compartment.id == 'c':
                volumes[species.id] = self.volume
            else:
                volumes[species.id] = self.extracellularVolume
        return volumes

    @staticmethod
    def calcReactionRates(reactions, speciesConcentrations):
        # calculate reaction rates
        rates = np.full(len(reactions), np.nan)
        for iRxn, rxn in enumerate(reactions):
            if rxn.rateLaw:
                rates[iRxn] = eval(rxn.rateLaw.transcoded, {}, {
                                   'speciesConcentrations': speciesConcentrations, 'Vmax': rxn.vmax, 'Km': rxn.km})
        return rates

    @staticmethod
    def executeReaction(speciesCounts, reaction):
        # update species counts based on a reaction
        for part in reaction.participants:
            speciesCounts[part.id] += part.coefficient
        return speciesCounts

    def getComponentById(self, id, components):
        for component in components:
            if component.id == id:
                return component


class FbaSubmodel(Submodel):
    # Represents an FBA submodel

    metabolismProductionReaction = None
    exchangedSpecies = None

    cobraModel = None
    thermodynamicBounds = None
    exchangeRateBounds = None

    defaultFbaBound = 1e15

    dryWeight = np.nan
    reactionFluxes = np.zeros(0)
    growth = np.nan

    solver = 'glpk'

    def __init__(self, *args, **kwargs):
        Submodel.__init__(self, *args, **kwargs)
        self.algorithm = 'FBA'

    def setupSimulation(self):
        '''setup reaction participant, enzyme counts matrices'''
        Submodel.setupSimulation(self)

        '''Setup FBA'''
        cobraModel = CobraModel(self.id)
        self.cobraModel = cobraModel

        # setup metabolites
        cbMets = []
        for species in self.species:
            cbMets.append(CobraMetabolite(id=species.id, name=species.name))
        cobraModel.add_metabolites(cbMets)

        # setup reactions
        for rxn in self.reactions:
            cbRxn = CobraReaction(
                id=rxn.id,
                name=rxn.name,
                lower_bound=-self.defaultFbaBound if rxn.reversible else 0,
                upper_bound=self.defaultFbaBound,
            )
            cobraModel.add_reactions([cbRxn])

            cbMets = {}
            for part in rxn.participants:
                cbMets[part.id] = part.coefficient
                if rxn.id == 'MetabolismProduction' and part.id == 'H2O[c]' and self.solver == 'glpk':  # to compensate for GLPK bug
                    del cbMets[part.id]
            cbRxn.add_metabolites(cbMets)

        # add external exchange reactions
        self.exchangedSpecies = []
        for species in self.species:
            if species.compartment.id == 'e':
                cbRxn = CobraReaction(
                    id='%sEx' % species.species.id,
                    name='%s exchange' % species.species.name,
                    lower_bound=-self.defaultFbaBound,
                    upper_bound=self.defaultFbaBound,
                )
                cobraModel.add_reactions([cbRxn])
                cbRxn.add_metabolites({species.id: 1})

                self.exchangedSpecies.append(ExchangedSpecies(id=species.id, reactionIndex=cobraModel.reactions.index(cbRxn)))

        # add biomass exchange reaction
        cbRxn = CobraReaction(
            id='BiomassEx',
            name='Biomass exchange',
            lower_bound=0,
            upper_bound=self.defaultFbaBound,
        )
        cobraModel.add_reactions([cbRxn])
        cbRxn.add_metabolites({'Biomass[c]': -1})

        '''Bounds'''
        # thermodynamic
        lower_bounds = []
        upper_bounds = []
        for rxn in cobraModel.reactions:
            lower_bounds.append(rxn.lower_bound)
            upper_bounds.append(rxn.upper_bound)
        self.thermodynamicBounds = {
            'lower': np.array(lower_bounds),
            'upper': np.array(upper_bounds),
        }

        # exchange reactions
        carbonExRate = self.getComponentById('carbonExchangeRate', self.parameters).value
        nonCarbonExRate = self.getComponentById('nonCarbonExchangeRate', self.parameters).value
        self.exchangeRateBounds = {
            'lower': np.full(len(cobraModel.reactions), -np.nan),
            'upper': np.full(len(cobraModel.reactions),  np.nan),
        }
        for exSpecies in self.exchangedSpecies:
            if self.getComponentById(exSpecies.id, self.species).species.containsCarbon():
                self.exchangeRateBounds['lower'][exSpecies.reactionIndex] = -carbonExRate
                self.exchangeRateBounds['upper'][exSpecies.reactionIndex] = carbonExRate
            else:
                self.exchangeRateBounds['lower'][exSpecies.reactionIndex] = -nonCarbonExRate
                self.exchangeRateBounds['upper'][exSpecies.reactionIndex] = nonCarbonExRate

        '''Setup reactions'''
        self.metabolismProductionReaction = {
            'index': cobraModel.reactions.index(cobraModel.reactions.get_by_id('MetabolismProduction')),
            'reaction': self.getComponentById('MetabolismProduction', self.reactions),
        }

        cobraModel.objective = 'MetabolismProduction'
        cobraModel.solver = self.solver

    def updateLocalCellState(self, model):
        Submodel.updateLocalCellState(self, model)
        self.dryWeight = model.dryWeight

    def updateGlobalCellState(self, model):
        Submodel.updateGlobalCellState(self, model)
        model.growth = self.growth

    def calcReactionFluxes(self, timeStep=1):
        '''calculate growth rate'''
        solution = self.cobraModel.optimize()
        assert(solution.status == 'optimal')

        self.reactionFluxes = solution.fluxes
        self.growth = self.reactionFluxes[self.metabolismProductionReaction['index']]  # fraction cell/s

    def updateMetabolites(self, timeStep=1):
        # biomass production
        for part in self.metabolismProductionReaction['reaction'].participants:
            self.speciesCounts[part.id] -= self.growth * part.coefficient * timeStep

        # external nutrients
        for exSpecies in self.exchangedSpecies:
            self.speciesCounts[exSpecies.id] += self.reactionFluxes[exSpecies.reactionIndex] * timeStep

    def calcReactionBounds(self,  timeStep=1):
        # thermodynamics
        lowerBounds = self.thermodynamicBounds['lower'].copy()
        upperBounds = self.thermodynamicBounds['upper'].copy()

        # rate laws
        upperBounds[0:len(self.reactions)] = util.nanminimum(
            upperBounds[0:len(self.reactions)],
            self.calcReactionRates(self.reactions, self.getSpeciesConcentrations()) * self.volume * N_AVOGADRO,
        )

        # external nutrients availability
        for exSpecies in self.exchangedSpecies:
            upperBounds[exSpecies.reactionIndex] = max(0, np.minimum(
                upperBounds[exSpecies.reactionIndex], self.speciesCounts[exSpecies.id]) / timeStep)

        # exchange bounds
        lowerBounds = util.nanmaximum(lowerBounds, self.dryWeight / 3600 * N_AVOGADRO * 1e-3 * self.exchangeRateBounds['lower'])
        upperBounds = util.nanminimum(upperBounds, self.dryWeight / 3600 * N_AVOGADRO * 1e-3 * self.exchangeRateBounds['upper'])

        # return
        for i_rxn, rxn in enumerate(self.cobraModel.reactions):
            rxn.lower_bound = lowerBounds[i_rxn]
            rxn.upper_bound = upperBounds[i_rxn]


class SsaSubmodel(Submodel):
    # Represents an SSA submodel

    def __init__(self, *args, **kwargs):
        Submodel.__init__(self, *args, **kwargs)
        self.algorithm = 'SSA'

    def setupSimulation(self):
        Submodel.setupSimulation(self)


class Compartment(object):
    # Represents a compartment

    index = None
    id = ''
    name = ''
    initialVolume = None
    comments = ''

    def __init__(self, id='', name='', initialVolume=None, comments=''):
        self.id = id
        self.name = name
        self.initialVolume = initialVolume
        self.comments = comments


class Species(object):
    # Represents a species

    index = None
    id = ''
    name = ''
    structure = ''
    empiricalFormula = ''
    molecularWeight = None
    charge = None
    type = ''
    concentrations = []
    crossRefs = []
    comments = ''

    # fix
    def __init__(self, id='', name='', structure='', empiricalFormula='', molecularWeight=None,
                 charge=None, type='', concentrations=[], crossRefs=[], comments=''):

        self.id = id
        self.name = name
        self.structure = structure
        self.empiricalFormula = empiricalFormula
        self.molecularWeight = molecularWeight
        self.charge = charge
        self.type = type
        self.concentrations = concentrations
        self.crossRefs = crossRefs

    def containsCarbon(self):
        if self.empiricalFormula:
            return self.empiricalFormula.upper().find('C') != -1
        return False


class Reaction(object):
    # Represents a reaction

    index = None
    id = ''
    name = ''
    submodel = ''
    reversible = None
    participants = []
    enzyme = ''
    rateLaw = None
    vmax = None
    km = None
    crossRefs = []
    comments = ''

    # fix
    def __init__(self, id='', name='', submodel='', reversible=None, participants=[],
                 enzyme='', rateLaw='', vmax=None, km=None, crossRefs=[], comments=''):

        if vmax:
            vmax = float(vmax)
        if km:
            km = float(km)

        self.id = id
        self.name = name
        self.submodel = submodel
        self.reversible = reversible
        self.participants = participants
        self.enzyme = enzyme
        self.rateLaw = rateLaw
        self.vmax = vmax
        self.km = km
        self.crossRefs = crossRefs
        self.comments = comments


class Parameter(object):
    # Represents a model parameter

    index = None
    id = ''
    name = ''
    submodel = None
    value = None
    units = ''
    comments = ''

    def __init__(self, id='', name='', submodel='', value=None, units='', comments=''):
        self.id = id
        self.name = name
        self.submodel = submodel
        self.value = value
        self.units = units
        self.comments = comments


class Reference(object):
    # Represents a reference

    index = None
    id = ''
    name = ''
    crossRefs = []
    comments = ''

    # fix
    def __init__(self, id='', name='', crossRefs=[], comments=''):
        self.id = id
        self.name = name
        self.crossRefs = crossRefs
        self.comments = comments


class Concentration(object):
    # Represents a concentration in a compartment

    compartment = ''
    value = None

    def __init__(self, compartment='', value=None):
        self.compartment = compartment
        self.value = value


class SpeciesCompartment(object):
    # Represents a participant in a submodel

    index = None
    species = ''
    compartment = ''

    id = ''
    name = ''

    def __init__(self, index=None, species='', compartment=''):
        self.index = index
        self.species = species
        self.compartment = compartment

    def calcIdName(self):
        self.id = '%s[%s]' % (self.species.id, self.compartment.id)
        self.name = '%s (%s)' % (self.species.name, self.compartment.name)


class ExchangedSpecies(object):
    # Represents an external

    id = ''
    reactionIndex = None

    def __init__(self, id='', reactionIndex=None):
        self.id = id
        self.reactionIndex = reactionIndex


class ReactionParticipant(object):
    # Represents a participant in a reaction

    species = ''
    compartment = ''
    coefficient = None

    id = ''
    name = ''

    def __init__(self, species='', compartment='', coefficient=None):
        self.species = species
        self.compartment = compartment
        self.coefficient = coefficient

    def calcIdName(self):
        self.id = '%s[%s]' % (self.species.id, self.compartment.id)
        self.name = '%s (%s)' % (self.species.name, self.compartment.name)


class RateLaw(object):
    # Represents a rate law

    native = ''
    transcoded = ''

    def __init__(self, native=''):
        self.native = native or ''

    def getModifiers(self, species, compartments):
        # get modifiers of rate law
        modifiers = []
        for spec in species:
            for comp in compartments:
                id = '%s[%s]' % (spec.id, comp.id)
                if self.native.find(id) != -1:
                    modifiers.append(id)
        return modifiers

    def transcode(self, species, compartments):
        # transcoded for python
        self.transcoded = self.native

        for spec in species:
            for comp in compartments:
                id = '%s[%s]' % (spec.id, comp.id)
                self.transcoded = self.transcoded.replace(id, "speciesConcentrations['%s']" % id)


class DatabaseReference(object):
    # Represents a database reference to an external database

    source = ''
    id = ''

    def __init__(self, source='', id=''):
        self.source = source
        self.id = id


def getModelFromExcel(filename):
    # Reads model from Excel file into a Python object

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", "Discarded range with reserved name", UserWarning)
        wb = load_workbook(filename=filename)

    # initialize model object
    model = Model()

    '''Read details from Excel'''
    # submodels
    ws = wb['Submodels']
    for iRow in range(2, ws.max_row + 1):
        id = str(ws.cell(row=iRow, column=1).value)
        name = ws.cell(row=iRow, column=2).value
        algorithm = ws.cell(row=iRow, column=3).value
        if algorithm == 'FBA':
            subModel = FbaSubmodel(id=id, name=name)
        elif algorithm == 'SSA':
            subModel = SsaSubmodel(id=id, name=name)
        model.submodels.append(subModel)

    # compartments
    ws = wb['Compartments']
    for iRow in range(2, ws.max_row + 1):
        model.compartments.append(Compartment(
            id=str(ws.cell(row=iRow, column=1).value),
            name=ws.cell(row=iRow, column=2).value,
            initialVolume=float(ws.cell(row=iRow, column=3).value),
            comments=ws.cell(row=iRow, column=4).value,
        ))

    # species
    ws = wb['Species']
    for iRow in range(2, ws.max_row + 1):
        mwStr = ws.cell(row=iRow, column=5).value
        if mwStr:
            mw = float(mwStr)
        else:
            mw = None

        chargeStr = ws.cell(row=iRow, column=6).value
        if chargeStr:
            charge = float(chargeStr)
        else:
            charge = None

        model.species.append(Species(
            id=str(ws.cell(row=iRow, column=1).value),
            name=ws.cell(row=iRow, column=2).value,
            structure=ws.cell(row=iRow, column=3).value,
            empiricalFormula=ws.cell(row=iRow, column=4).value,
            molecularWeight=mw,
            charge=charge,
            type=ws.cell(row=iRow, column=7).value,
            concentrations=[
                Concentration(compartment='c', value=float(ws.cell(row=iRow, column=8).value or 0)),
                Concentration(compartment='e', value=float(ws.cell(row=iRow, column=9).value or 0)),
            ],
            crossRefs=[
                DatabaseReference(
                    source=ws.cell(row=iRow, column=10).value,
                    id=ws.cell(row=iRow, column=11).value,
                ),
            ],
            comments=ws.cell(row=iRow, column=12).value,
        ))

    # reactions
    ws = wb['Reactions']

    for iRow in range(2, ws.max_row + 1):
        stoichiometry = parseStoichiometry(ws.cell(row=iRow, column=4).value)

        rateLawStr = ws.cell(row=iRow, column=6).value
        if rateLawStr:
            rateLaw = RateLaw(rateLawStr)
        else:
            rateLaw = None

        model.reactions.append(Reaction(
            id=str(ws.cell(row=iRow, column=1).value),
            name=ws.cell(row=iRow, column=2).value,
            submodel=ws.cell(row=iRow, column=3).value,
            reversible=stoichiometry['reversible'],
            participants=stoichiometry['participants'],
            enzyme=ws.cell(row=iRow, column=5).value,
            rateLaw=rateLaw,
            vmax=ws.cell(row=iRow, column=7).value,
            km=ws.cell(row=iRow, column=8).value,
            crossRefs=[
                DatabaseReference(
                    source=ws.cell(row=iRow, column=9).value,
                    id=ws.cell(row=iRow, column=10).value,
                ),
            ],
            comments=ws.cell(row=iRow, column=11).value,
        ))

    # parameters
    ws = wb['Parameters']
    for iRow in range(2, ws.max_row + 1):
        model.parameters.append(Parameter(
            id=str(ws.cell(row=iRow, column=1).value),
            name=ws.cell(row=iRow, column=2).value,
            submodel=ws.cell(row=iRow, column=3).value,
            value=float(ws.cell(row=iRow, column=4).value),
            units=ws.cell(row=iRow, column=5).value,
            comments=ws.cell(row=iRow, column=6).value,
        ))

    # references
    ws = wb['References']
    for iRow in range(2, ws.max_row + 1):
        model.references.append(Reference(
            id=str(ws.cell(row=iRow, column=1).value),
            name=ws.cell(row=iRow, column=2).value,
            crossRefs=[
                DatabaseReference(
                    source=ws.cell(row=iRow, column=3).value,
                    id=ws.cell(row=iRow, column=4).value,
                ),
            ],
            comments=ws.cell(row=iRow, column=5).value,
        ))

    '''set component indices'''
    model.setComponentIndices()

    '''deserialize references'''
    # species concentration
    for species in model.species:
        for conc in species.concentrations:
            id = conc.compartment
            obj = model.getComponentById(id, model.compartments)
            conc.compartment = obj

    # reaction submodel, participant species, participant compartments, enzymes
    for reaction in model.reactions:
        id = reaction.submodel
        obj = model.getComponentById(id, model.submodels)
        reaction.submodel = obj

        for part in reaction.participants:

            id = part.species
            obj = model.getComponentById(id, model.species)
            part.species = obj

            id = part.compartment
            obj = model.getComponentById(id, model.compartments)
            part.compartment = obj

            part.calcIdName()

        id = reaction.enzyme
        obj = model.getComponentById(id, model.species)
        reaction.enzyme = obj

    # parameter submodels
    for param in model.parameters:
        id = param.submodel
        if id:
            obj = model.getComponentById(id, model.submodels)
            param.submodel = obj

    ''' Assemble back references'''
    for subModel in model.submodels:
        subModel.reactions = []
        subModel.species = []
        subModel.parameters = []
    for rxn in model.reactions:
        rxn.submodel.reactions.append(rxn)
        for part in rxn.participants:
            rxn.submodel.species.append('%s[%s]' % (part.species.id, part.compartment.id))
        if rxn.enzyme:
            rxn.submodel.species.append('%s[%s]' % (rxn.enzyme.id, 'c'))
        if rxn.rateLaw:
            rxn.submodel.species += rxn.rateLaw.getModifiers(model.species, model.compartments)

    for param in model.parameters:
        if param.submodel:
            param.submodel.parameters.append(param)

    for subModel in model.submodels:
        speciesStrArr = list(set(subModel.species))
        speciesStrArr.sort()
        subModel.species = []
        for index, speciesStr in enumerate(speciesStrArr):
            speciesId, compId = speciesStr.split('[')
            compId = compId[0:-1]
            speciesComp = SpeciesCompartment(
                index=index,
                species=model.getComponentById(speciesId, model.species),
                compartment=model.getComponentById(compId, model.compartments),
            )
            speciesComp.calcIdName()
            subModel.species.append(speciesComp)

    '''Transcode rate laws'''
    for rxn in model.reactions:
        if rxn.rateLaw:
            rxn.rateLaw.transcode(model.species, model.compartments)

    '''Prepare submodels for computation'''
    model.setupSimulation()

    '''Return'''
    return model


def parseStoichiometry(rxnStr):
    # Parse a string representing the stoichiometry of a reaction into a Python object

    # Split stoichiometry in to global compartment, left-hand side, right-hand side, reversibility indictor
    rxnMatch = re.match('(?P<compartment>\[([a-z])\]: )?(?P<lhs>((\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])? \+ )*(\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])?) (?P<direction>[<]?)==> (?P<rhs>((\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])? \+ )*(\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])?)', rxnStr, flags=re.I)
    if rxnMatch is None:
        raise ValueError('Invalid stoichiometry: %s' % rxnStr)

    # Determine reversiblity
    rxnDict = rxnMatch.groupdict()
    reversible = rxnDict['direction'] == '<'

    # Determine if global compartment for reaction was specified
    if rxnDict['compartment'] is None:
        globalComp = None
    else:
        globalComp = re.match('\[(?P<compartment>[a-z])\]', rxnDict['compartment'], flags=re.I).groupdict()['compartment']

    # initialize array of reaction participants
    participants = []

    # Parse left-hand side
    for rxnPartStr in rxnDict['lhs'].split(' + '):
        rxnPartDict = re.match(
            '(\((?P<coefficient>\d*\.?\d*([e][-+]?[0-9]+)?)\) )?(?P<species>[a-z0-9\-_]+)(\[(?P<compartment>[a-z])\])?', rxnPartStr, flags=re.I).groupdict()

        species = rxnPartDict['species']
        compartment = rxnPartDict['compartment'] or globalComp
        coefficient = float(rxnPartDict['coefficient'] or 1)

        participants.append(ReactionParticipant(
            species=species,
            compartment=compartment,
            coefficient=-coefficient,
        ))

    # Parse right-hand side
    for rxnPartStr in rxnDict['rhs'].split(' + '):
        rxnPartDict = re.match(
            '(\((?P<coefficient>\d*\.?\d*([e][-+]?[0-9]+)?)\) )?(?P<species>[a-z0-9\-_]+)(\[(?P<compartment>[a-z])\])?', rxnPartStr, flags=re.I).groupdict()

        species = rxnPartDict['species']
        compartment = rxnPartDict['compartment'] or globalComp
        coefficient = float(rxnPartDict['coefficient'] or 1)

        participants.append(ReactionParticipant(
            species=species,
            compartment=compartment,
            coefficient=coefficient,
        ))

    return {
        'reversible': reversible,
        'participants': participants,
    }
